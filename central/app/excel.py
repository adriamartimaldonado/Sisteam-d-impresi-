# -*- coding: utf-8 -*-
"""Lectura y validacion del Excel de un pedido (§14 del documento de contexto).

Reglas que importan:
- Se lee TODO como texto y se mira el TIPO de cada celda. Si una celda de un campo
  de codigo/GTIN/EPC viene como numero, Excel pudo comerse los ceros a la izquierda
  o pasar el numero a notacion cientifica: eso es una tirada a la basura, asi que se
  rechaza indicando la fila.
- Se valida el fichero ENTERO. Si algo falla, se rechaza el pedido completo con la
  lista de (fila, columna, motivo). Nunca se imprime medio Excel.
- Se comprueban duplicados dentro del propio fichero, en especial en EPC.

`validar(...)` es una funcion pura (no toca red ni BD): facil de testear.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


# ------------------------------------------------------------------ tipos
@dataclass(frozen=True)
class Campo:
    """Especificacion de una columna esperada (viene de plantilla.campos)."""
    nombre: str
    requerido: bool = True
    tipo: str = "texto"          # texto | gtin | epc | entero
    unico: bool = False

    @staticmethod
    def desde_dict(d: dict[str, Any]) -> "Campo":
        return Campo(
            nombre=str(d["nombre"]),
            requerido=bool(d.get("requerido", True)),
            tipo=str(d.get("tipo", "texto")).lower(),
            unico=bool(d.get("unico", False)),
        )


@dataclass(frozen=True)
class Celda:
    valor: str
    numerica: bool


@dataclass(frozen=True)
class ErrorFila:
    fila: int | None       # numero de fila de Excel (cabecera = 1, datos desde 2); None = global
    columna: str | None
    motivo: str

    def como_dict(self) -> dict[str, Any]:
        return {"fila": self.fila, "columna": self.columna, "motivo": self.motivo}


class ExcelInvalido(Exception):
    """El Excel no se pudo ni abrir/leer (fichero corrupto, hoja inexistente...)."""


# ------------------------------------------------------------- lectura
def _celda(valor: Any, tipo_dato: str | None) -> Celda:
    if valor is None:
        return Celda("", False)
    numerica = tipo_dato == "n"
    if isinstance(valor, float) and valor.is_integer():
        texto = str(int(valor))          # 12345.0 -> "12345"
    else:
        texto = str(valor)
    return Celda(texto.strip(), numerica)


def abrir_filas(contenido: bytes, hoja: str | None = None
                ) -> tuple[list[str], list[list[Celda]]]:
    """Devuelve (cabeceras, filas). La primera fila es la cabecera. Cada celda lleva
    su valor como texto y si Excel la guardo como numero."""
    try:
        wb = load_workbook(BytesIO(contenido), data_only=True, read_only=True)
    except Exception as exc:   # openpyxl lanza tipos variados; se normaliza
        raise ExcelInvalido(f"No se pudo abrir el Excel: {exc}") from exc
    try:
        ws = wb[hoja] if hoja else wb[wb.sheetnames[0]]
    except KeyError as exc:
        raise ExcelInvalido(f"La hoja '{hoja}' no existe en el Excel.") from exc

    filas_iter = ws.iter_rows()
    try:
        primera = next(filas_iter)
    except StopIteration:
        wb.close()
        raise ExcelInvalido("El Excel esta vacio.")

    cabeceras = [_celda(c.value, c.data_type).valor for c in primera]
    filas: list[list[Celda]] = []
    for fila in filas_iter:
        celdas = [_celda(c.value, c.data_type) for c in fila]
        # ignorar filas totalmente vacias (Excel suele arrastrar filas fantasma)
        if any(c.valor for c in celdas):
            filas.append(celdas)
    wb.close()
    return cabeceras, filas


# ----------------------------------------------------------- validacion
def _digito_control_gtin(cuerpo: str) -> str:
    """Digito de control GS1 (mod 10) del cuerpo (sin el ultimo digito)."""
    suma = 0
    for i, ch in enumerate(reversed(cuerpo)):
        n = int(ch)
        suma += n * (3 if i % 2 == 0 else 1)
    return str((10 - (suma % 10)) % 10)


def _valida_gtin(v: str) -> str | None:
    if not v.isdigit():
        return "GTIN no numerico"
    if len(v) not in (8, 12, 13, 14):
        return f"GTIN de longitud {len(v)} (se esperan 8, 12, 13 o 14)"
    if _digito_control_gtin(v[:-1]) != v[-1]:
        return "digito de control de GTIN incorrecto"
    return None


def _valida_epc(v: str) -> str | None:
    u = v.upper()
    if any(c not in "0123456789ABCDEF" for c in u):
        return "EPC no hexadecimal"
    if len(u) % 4 != 0:
        return f"EPC de longitud {len(u)} (debe ser multiplo de 4)"
    if not (16 <= len(u) <= 124):
        return f"EPC de longitud {len(u)} fuera de rango (16-124)"
    return None


def _valida_valor(celda: Celda, campo: Campo) -> str | None:
    """Devuelve el motivo de error del valor de una celda, o None si es valido.
    (No cubre 'vacio' ni 'numerica': eso se decide fuera con el contexto del campo.)"""
    v = celda.valor
    if campo.tipo == "gtin":
        return _valida_gtin(v)
    if campo.tipo == "epc":
        return _valida_epc(v)
    if campo.tipo == "entero":
        return None if v.isdigit() else "no es un entero"
    return None   # texto: cualquier cosa no vacia vale


def validar(cabeceras: list[str], filas: list[list[Celda]], campos: list[Campo]
            ) -> tuple[list[dict], list[ErrorFila]]:
    """Valida las filas contra los campos esperados.

    Devuelve (datos, errores):
    - si `errores` no esta vacio, el pedido se rechaza ENTERO y `datos` va vacio;
    - si esta vacio, `datos` es la lista de dicts {campo: valor} por fila, en orden.
    """
    errores: list[ErrorFila] = []

    # Mapa cabecera(normalizada) -> indice de columna.
    indice: dict[str, int] = {}
    for i, cab in enumerate(cabeceras):
        clave = cab.strip().lower()
        if clave and clave not in indice:
            indice[clave] = i

    # 1) Todas las columnas requeridas deben existir. Si falta una, error global.
    col_de_campo: dict[str, int] = {}
    for campo in campos:
        idx = indice.get(campo.nombre.strip().lower())
        if idx is None:
            if campo.requerido:
                errores.append(ErrorFila(1, campo.nombre, "falta la columna requerida"))
        else:
            col_de_campo[campo.nombre] = idx
    if errores:
        return [], errores   # sin columnas no tiene sentido validar filas

    # 2) Validar cada fila. Excel: cabecera = fila 1, primeros datos = fila 2.
    vistos: dict[str, dict[str, int]] = {c.nombre: {} for c in campos if c.unico}
    datos: list[dict] = []
    for n, fila in enumerate(filas, start=2):
        registro: dict[str, str] = {}
        for campo in campos:
            idx = col_de_campo.get(campo.nombre)
            celda = fila[idx] if (idx is not None and idx < len(fila)) else Celda("", False)

            if not celda.valor:
                if campo.requerido:
                    errores.append(ErrorFila(n, campo.nombre, "vacio"))
                registro[campo.nombre] = ""
                continue

            # Celda numerica en un campo de codigo/GTIN/EPC: Excel pudo mangear el dato.
            if celda.numerica and campo.tipo in ("texto", "gtin", "epc"):
                errores.append(ErrorFila(
                    n, campo.nombre,
                    "celda numerica: Excel pudo perder ceros a la izquierda o pasar a "
                    "notacion cientifica (guarda la columna como texto)"))
                registro[campo.nombre] = celda.valor
                continue

            motivo = _valida_valor(celda, campo)
            if motivo:
                errores.append(ErrorFila(n, campo.nombre, motivo))
                registro[campo.nombre] = celda.valor
                continue

            # Duplicado dentro del propio fichero (sobre todo EPC).
            if campo.unico:
                previa = vistos[campo.nombre].get(celda.valor)
                if previa is not None:
                    errores.append(ErrorFila(
                        n, campo.nombre, f"duplicado (ya aparece en la fila {previa})"))
                else:
                    vistos[campo.nombre][celda.valor] = n

            registro[campo.nombre] = celda.valor
        datos.append(registro)

    if errores:
        return [], errores
    return datos, []
